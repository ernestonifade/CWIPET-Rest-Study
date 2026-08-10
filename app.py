import os
import uuid
import openpyxl
import pandas as pd
import streamlit as st
from figures.figure1 import load_fig1_results, render_figure1
from figures.figure2 import load_fig2_results, render_figure2
from figures.figure3 import load_fig3_results, render_figure3
from figures.figure4 import load_fig4_results, render_figure4
from figures.figure5 import load_fig5_results, render_figure5
from figures.figure6 import load_fig6_results, render_figure6
from utils import render_searchable_table

# --- 1. STREAMLIT PAGE CONFIGURATION (WIDE & OPEN) ---
st.set_page_config(
    page_title="CWIPET Rest Study Dashboard",
    page_icon="🧬",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
    <style>
    /* Force top-alignment and scale radio buttons for high visibility */
    .stRadio div[role="radiogroup"] {
        gap: 10px;
    }
    .stRadio div[role="radiogroup"] label {
        align-items: flex-start !important;
        cursor: pointer;
    }
    .stRadio div[role="radiogroup"] label div:first-child {
        margin-top: 3px;
    }
    /* Scale up the radio inputs to make them larger and easier to click */
    .stRadio div[role="radiogroup"] input[type="radio"] {
        transform: scale(1.0);
        accent-color: #1f77b4;
        margin-right: 6px;
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
<style>
    .block-container {
        padding-top: 1.5rem;
        padding-bottom: 2rem;
        max-width: 98%;
    }
    h1 {
        font-size: 1.8rem !important;
        margin-bottom: 0px !important;
    }
    p {
        font-size: 0.95rem;
    }
    .stDownloadButton button {
        background-color: #28a745 !important;
        color: white !important;
        font-weight: bold;
        width: 100%;
    }
</style>
""",
    unsafe_allow_html=True,
)

st.title("🧬 Cold-Water Immersion Rest Study")
st.caption("Interactive Manuscript Dashboard & Statistical Summary")
st.markdown("---")

# --- 2. SIDEBAR NAVIGATION ---
st.sidebar.header("📌 Navigation")
selected_figure = st.sidebar.radio(
    "Select Figure:",
    [
        "Figure 1: Metabolite and Cytokine responses to different degrees of CWI",
        "Figure 2: Metabolite and Cytokine correlations in response to CWI",
        "Figure 3: Pathway enrichment correlating cytokines and metabolites",
        "Figure 4: Body temperatures responses to different degrees of CWI",
        "Figure 5: Metabolite response to CWI given bodymetrics",
        "Figure 6: Cytokine response to CWI given bodymetrics and bodytemperatures",
        
        
    ],
    index=0,
)

st.sidebar.markdown("---")

# Helper function to write DataFrames safely using openpyxl
def export_sheets_to_excel(filename, sheets_dict):
  wb = openpyxl.Workbook()
  default_sheet = wb.active
  wb.remove(default_sheet)

  for sheet_name, df in sheets_dict.items():
    if df is not None and not df.empty:
      ws = wb.create_sheet(title=sheet_name)
      ws.append(list(df.columns))
      for row in df.itertuples(index=False, name=None):
        ws.append(list(row))

  if len(wb.worksheets) == 0:
    ws = wb.create_sheet(title="Info")
    ws.append(["Note"])
    ws.append(["No data available"])

  wb.save(filename)


# --- 3. PAGE ROUTING & RENDER CALLS ---
if selected_figure == "Figure 1: Metabolite and Cytokine responses to different degrees of CWI":
  render_figure1()
elif (
    selected_figure
    == "Figure 2: Metabolite and Cytokine correlations in response to CWI"
):
  render_figure2()
elif selected_figure == "Figure 3: Pathway enrichment correlating cytokines and metabolites":
  render_figure3()
elif selected_figure == "Figure 4: Body temperatures responses to different degrees of CWI":
  render_figure4()
elif selected_figure == "Figure 5: Metabolite response to CWI given bodymetrics":
  render_figure5()
elif selected_figure == "Figure 6: Cytokine response to CWI given bodymetrics and bodytemperatures":
  render_figure6()



# --- 4. ONE-CLICK INSTANT EXPORT HANDLER ---
st.sidebar.header("📥 Export Statistical Reports")

if (
    selected_figure
    == "Figure 1: Metabolite and Cytokine responses to different degrees of CWI"
):
  ancova_df, posthoc_df, long_df = load_fig1_results()
  out_name = "Figure1_EV_Full_Stats_Report.xlsx"
  export_sheets_to_excel(
      out_name,
      {
          "RM_ANCOVA_Stats": ancova_df,
          "PostHoc_Contrasts": posthoc_df,
          "Raw_Data": long_df,
      },
  )

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 1 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig1_{uuid.uuid4()}",
    )

elif selected_figure == "Figure 2: Metabolite and Cytokine correlations in response to CWI":
  data = load_fig2_results()
  out_name = "Figure2_Metabolite_Cytokine_Corr_Report.xlsx"
  export_sheets_to_excel(
      out_name,
      {
          "Cyto_Metabolite_RM_Corr": data.get("cyto_rm"),
          "Cyto_Metabolite_Baseline": data.get("cyto_baseline"),
          "Cyto_Metabolite_Delta_Windows": data.get("cyto_delta"),
          
      },
  )

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 2 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig2_{uuid.uuid4()}",
    )

elif (
    selected_figure
    == "Figure 3: Pathway enrichment correlating cytokines and metabolites"
):
  def find_pathway_file(candidates):
    for path in candidates:
      if os.path.exists(path):
        return path
    return None

  out_name = "Pathway enrichment correlating cytokines and metabolites.xlsx"
  prot_path = find_pathway_file([
      "data/enrichment_results_for_correlating_metabolites.csv",
      "../data/enrichment_results_for_correlating_metabolites.csv",
      "enrichment_results_for_correlating_metabolites.csv",
  ])
  cyt_path = find_pathway_file([
      "data/enrichment_results_for_correlating_cytokines.csv",
      "../data/enrichment_results_for_correlating_cytokines.csv",
      "enrichment_results_for_correlating_cytokines.csv",
  ])

  sheets_data = {}
  if prot_path and os.path.exists(prot_path):
    df_prot = pd.read_csv(prot_path)
    if not df_prot.empty:
      sheets_data["Metabolite_Pathways"] = df_prot

  if cyt_path and os.path.exists(cyt_path):
    df_cyt = pd.read_csv(cyt_path)
    if not df_cyt.empty:
      sheets_data["Cytokine_Pathways"] = df_cyt

  export_sheets_to_excel(out_name, sheets_data)

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 3 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig7_{uuid.uuid4()}",
    )

elif (
    selected_figure
    == "Figure 4: Body temperatures responses to different degrees of CWI"
):
  ancova_df, posthoc_df, long_df = load_fig4_results()
  out_name = "Figure4_Full_Stats_Report.xlsx"
  export_sheets_to_excel(
      out_name,
      {
          "RM_ANCOVA_Stats": ancova_df,
          "PostHoc_Contrasts": posthoc_df,
          "Raw_Data": long_df,
      },
  )

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 4 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig4_{uuid.uuid4()}",
    )

elif selected_figure == "Figure 5: Metabolite response to CWI given bodymetrics":
  data = load_fig5_results()
  out_name = "Figure5_Metabolite response to CWI given bodymetrics_Report.xlsx"
  export_sheets_to_excel(
      out_name,
      {
          "Metabolite response to CWI given bodymetrics": data.get("Bodymetric_Met"),
      },
  )

  with open(out_name, "rb") as f:
    st.sidebar.download_button(
        label="📥 Download Figure 5 Report (.xlsx)",
        data=f,
        file_name=out_name,
        mime=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        key=f"dl_fig5_{uuid.uuid4()}",
    )



# --- COMPLETE REPOSITORY DOWNLOAD (Placed right below the active page report button) ---
st.sidebar.markdown("---")
st.sidebar.markdown("### 🗂️ Complete Study Repository")

import io
import zipfile

zip_buffer = io.BytesIO()
data_folder = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

if os.path.exists(data_folder):
  with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zip_file:
    for root, dirs, files in os.walk(data_folder):
      for file in files:
        if file.lower().endswith(
            (".xlsx", ".xls", ".csv", ".txt", ".gmt", ".gmx")
        ):
          file_path = os.path.join(root, file)
          arcname = os.path.relpath(file_path, data_folder)
          zip_file.write(file_path, arcname=arcname)

  zip_buffer.seek(0)

  st.sidebar.download_button(
      label="📥 Download All Raw Datasets & GMTs (.zip)",
      data=zip_buffer,
      file_name="CWIPET_Rest_Study_Complete_Data.zip",
      mime="application/zip",
      key=f"dl_all_repo_data_{uuid.uuid4()}",
  )
            
    
